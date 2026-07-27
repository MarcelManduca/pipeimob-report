import os
import csv
import openpyxl
import hashlib
import json
from typing import List, Tuple, Dict, Any, Optional

MAX_FILE_SIZE_BYTES = int(os.getenv("IMPORT_MAX_FILE_SIZE_BYTES", 5 * 1024 * 1024))
MAX_SHEETS = int(os.getenv("IMPORT_MAX_SHEETS", 10))
MAX_ROWS = int(os.getenv("IMPORT_MAX_ROWS", 5000))
MAX_COLUMNS = int(os.getenv("IMPORT_MAX_COLUMNS", 100))

HEADER_ALIASES_CODIGO = {
    "cod. imóvel", "cod imóvel", "código imóvel", "código do imóvel", "codigo imovel"
}
HEADER_ALIASES_RESPONSAVEL = {
    "responsável", "responsavel"
}
HEADER_ALIASES_GERENTE = {
    "gerente"
}
HEADER_ALIASES_IMOVEL = {
    "imóvel", "imovel"
}
HEADER_ALIASES_CADASTRO = {
    "data de cadastro"
}
HEADER_ALIASES_ASSINATURA = {
    "data assinatura ccv", "data assin ccv"
}

def normalize_header(h: str) -> str:
    if not h:
        return ""
    # strip spaces, convert to lowercase, collapse spaces
    return " ".join(str(h).strip().lower().split())

def map_headers(headers: List[str]) -> Dict[str, Optional[int]]:
    mapping = {
        "codigo": None,
        "responsavel": None,
        "gerente": None,
        "imovel": None,
        "cadastro": None,
        "assinatura": None
    }
    for idx, h in enumerate(headers):
        norm = normalize_header(h)
        if norm in HEADER_ALIASES_CODIGO:
            mapping["codigo"] = idx
        elif norm in HEADER_ALIASES_RESPONSAVEL:
            mapping["responsavel"] = idx
        elif norm in HEADER_ALIASES_GERENTE:
            mapping["gerente"] = idx
        elif norm in HEADER_ALIASES_IMOVEL:
            mapping["imovel"] = idx
        elif norm in HEADER_ALIASES_CADASTRO:
            mapping["cadastro"] = idx
        elif norm in HEADER_ALIASES_ASSINATURA:
            mapping["assinatura"] = idx
    return mapping

def clean_cell_value(val: Any) -> Optional[str]:
    if val is None:
        return None
    
    # Handle openpyxl float integer check (e.g. 39177.0 -> "39177")
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)
        
    s = str(val).strip()
    if s == "":
        return None
    return s

class ContractsControlImportParser:
    @staticmethod
    def parse_file(file_path: str, filename: str) -> List[Dict[str, Any]]:
        # 1. Size Validation
        file_size = os.path.getsize(file_path)
        if file_size > MAX_FILE_SIZE_BYTES:
            raise ValueError(f"File size {file_size} exceeds limit of {MAX_FILE_SIZE_BYTES} bytes.")

        ext = os.path.splitext(filename)[1].lower()
        if ext not in (".xlsx", ".csv"):
            raise ValueError("Unsupported file format. Only .xlsx and .csv files are allowed.")

        rows = []
        if ext == ".xlsx":
            rows = ContractsControlImportParser._parse_xlsx(file_path)
        else:
            rows = ContractsControlImportParser._parse_csv(file_path)
            
        return rows

    @staticmethod
    def _parse_xlsx(file_path: str) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        
        # Validation: Max sheets
        if len(sheets) > MAX_SHEETS:
            raise ValueError(f"Number of sheets {len(sheets)} exceeds limit of {MAX_SHEETS}.")

        parsed_rows = []
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            
            # Validation: Max columns (approximate check by headers or first few rows)
            # Validation: Max rows check
            row_count = 0
            headers = []
            header_map = {}
            
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                row_count += 1
                if row_count > MAX_ROWS:
                    raise ValueError(f"Sheet '{sheet_name}' exceeds the row limit of {MAX_ROWS}.")
                
                # Check column count
                if len(row) > MAX_COLUMNS:
                    raise ValueError(f"Row {r_idx} in sheet '{sheet_name}' has {len(row)} columns, exceeding limit of {MAX_COLUMNS}.")
                
                if r_idx == 1:
                    headers = [str(x) if x is not None else "" for x in row]
                    header_map = map_headers(headers)
                    # If both codigo and responsavel are not mapped, this sheet might not be target
                    continue

                # Parse data row
                code_val = clean_cell_value(row[header_map["codigo"]]) if header_map["codigo"] is not None else None
                resp_val = clean_cell_value(row[header_map["responsavel"]]) if header_map["responsavel"] is not None else None
                gerente_val = clean_cell_value(row[header_map["gerente"]]) if header_map["gerente"] is not None else None
                imovel_val = clean_cell_value(row[header_map["imovel"]]) if header_map["imovel"] is not None else None
                cad_val = clean_cell_value(row[header_map["cadastro"]]) if header_map["cadastro"] is not None else None
                assin_val = clean_cell_value(row[header_map["assinatura"]]) if header_map["assinatura"] is not None else None

                # Skip completely empty rows
                if code_val is None and resp_val is None and gerente_val is None and imovel_val is None:
                    continue

                parsed_rows.append({
                    "aba": sheet_name,
                    "linha": r_idx,
                    "codigo_imovel": code_val,
                    "responsavel_planilha": resp_val,
                    "gerente": gerente_val,
                    "nome_imovel": imovel_val,
                    "data_cadastro": cad_val,
                    "data_assinatura_ccv": assin_val
                })
        return parsed_rows

    @staticmethod
    def _parse_csv(file_path: str) -> List[Dict[str, Any]]:
        # Read a sample to detect delimiter and validate row/column boundaries
        with open(file_path, mode="r", encoding="utf-8-sig", errors="ignore") as f:
            sample = f.read(4096)
            delim = ";" if ";" in sample else ","

        parsed_rows = []
        with open(file_path, mode="r", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=delim)
            row_count = 0
            header_map = {}
            
            for r_idx, row in enumerate(reader, start=1):
                row_count += 1
                if row_count > MAX_ROWS:
                    raise ValueError(f"CSV file exceeds the row limit of {MAX_ROWS}.")
                
                if len(row) > MAX_COLUMNS:
                    raise ValueError(f"Row {r_idx} in CSV has {len(row)} columns, exceeding limit of {MAX_COLUMNS}.")

                if r_idx == 1:
                    headers = [str(x).strip() for x in row]
                    header_map = map_headers(headers)
                    continue

                code_val = clean_cell_value(row[header_map["codigo"]]) if header_map["codigo"] is not None and header_map["codigo"] < len(row) else None
                resp_val = clean_cell_value(row[header_map["responsavel"]]) if header_map["responsavel"] is not None and header_map["responsavel"] < len(row) else None
                gerente_val = clean_cell_value(row[header_map["gerente"]]) if header_map["gerente"] is not None and header_map["gerente"] < len(row) else None
                imovel_val = clean_cell_value(row[header_map["imovel"]]) if header_map["imovel"] is not None and header_map["imovel"] < len(row) else None
                cad_val = clean_cell_value(row[header_map["cadastro"]]) if header_map["cadastro"] is not None and header_map["cadastro"] < len(row) else None
                assin_val = clean_cell_value(row[header_map["assinatura"]]) if header_map["assinatura"] is not None and header_map["assinatura"] < len(row) else None

                if code_val is None and resp_val is None and gerente_val is None and imovel_val is None:
                    continue

                parsed_rows.append({
                    "aba": "csv",
                    "linha": r_idx,
                    "codigo_imovel": code_val,
                    "responsavel_planilha": resp_val,
                    "gerente": gerente_val,
                    "nome_imovel": imovel_val,
                    "data_cadastro": cad_val,
                    "data_assinatura_ccv": assin_val
                })
        return parsed_rows
